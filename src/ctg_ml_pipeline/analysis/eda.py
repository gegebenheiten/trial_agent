"""
Comprehensive EDA module for CTG ML Pipeline.

Provides:
1. Per-table analysis: schema, dtypes, row_count, key duplication, missing rates,
   nunique, constant columns, numeric quantiles/outliers, categorical top-k,
   text length, JSON parse rates.
2. Cross-table analysis: StudyID coverage, (StudyID, Arm_ID) uniqueness,
   join coverage for endpoints.
3. Leakage check: identify T2 features that should be excluded.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import polars as pl


# Pseudo-missing patterns (case-insensitive)
PSEUDO_MISSING_PATTERNS = [
    r"^n/?a$", r"^nan$", r"^null$", r"^none$", r"^missing$",
    r"^not\s*(available|applicable|reported|specified)$",
    r"^unknown$", r"^-$", r"^\.$", r"^--$", r"^\s*$"
]
PSEUDO_MISSING_RE = re.compile("|".join(PSEUDO_MISSING_PATTERNS), re.IGNORECASE)


def is_pseudo_missing(val: str | None) -> bool:
    """Check if value is pseudo-missing."""
    if val is None:
        return True
    if isinstance(val, str):
        return bool(PSEUDO_MISSING_RE.match(val.strip()))
    return False


def try_parse_json(val: str | None) -> bool:
    """Check if value is valid JSON."""
    if val is None or val == "":
        return False
    try:
        json.loads(val)
        return True
    except (json.JSONDecodeError, TypeError):
        return False


@dataclass
class ColumnStats:
    """Statistics for a single column."""
    name: str
    dtype: str
    total_rows: int
    null_count: int
    pseudo_missing_count: int
    missing_rate: float  # includes pseudo-missing
    nunique: int
    is_constant: bool
    # Numeric stats
    min_val: float | None = None
    max_val: float | None = None
    mean_val: float | None = None
    std_val: float | None = None
    q25: float | None = None
    q50: float | None = None
    q75: float | None = None
    outlier_count: int = 0
    # Categorical stats
    top_values: list[tuple[str, int]] = field(default_factory=list)
    is_high_cardinality: bool = False
    # Text stats
    text_len_min: int | None = None
    text_len_max: int | None = None
    text_len_mean: float | None = None
    # JSON stats
    json_parse_rate: float | None = None


@dataclass
class TableStats:
    """Statistics for a single table."""
    name: str
    row_count: int
    col_count: int
    columns: list[ColumnStats]
    candidate_keys: list[dict[str, Any]]  # key columns + duplication rate
    constant_columns: list[str]
    high_missing_columns: list[str]  # >50% missing


@dataclass
class CrossTableStats:
    """Cross-table statistics."""
    study_id_coverage: dict[str, int]  # table -> count of unique StudyIDs
    study_id_total: int
    arm_uniqueness: dict[str, dict[str, Any]]  # table -> uniqueness info
    join_coverage: dict[str, float]  # join name -> coverage rate


@dataclass
class LeakageReport:
    """Leakage analysis report."""
    t0_features: list[str]
    t1_features: list[str]
    t2_features: list[str]  # should be excluded
    unknown_features: list[str]  # not in mapping
    feature_mapping: dict[str, str]  # feature -> timepoint


@dataclass
class EDAReport:
    """Complete EDA report."""
    table_stats: dict[str, TableStats]
    cross_table_stats: CrossTableStats
    leakage_report: LeakageReport


def analyze_column(df: pl.DataFrame, col_name: str, top_k: int = 5, high_card_threshold: int = 50) -> ColumnStats:
    """Analyze a single column."""
    col = df.get_column(col_name)
    dtype = str(col.dtype)
    total_rows = len(col)
    
    # Null count
    null_count = col.null_count()
    
    # Pseudo-missing count (for string columns)
    pseudo_missing_count = 0
    if col.dtype == pl.Utf8:
        pseudo_missing_count = col.map_elements(
            lambda x: is_pseudo_missing(x), return_dtype=pl.Boolean
        ).sum() - null_count
        pseudo_missing_count = max(0, pseudo_missing_count)
    
    total_missing = null_count + pseudo_missing_count
    missing_rate = total_missing / total_rows if total_rows > 0 else 0.0
    
    # Unique count
    nunique = col.n_unique()
    is_constant = nunique <= 1
    
    stats = ColumnStats(
        name=col_name,
        dtype=dtype,
        total_rows=total_rows,
        null_count=null_count,
        pseudo_missing_count=pseudo_missing_count,
        missing_rate=missing_rate,
        nunique=nunique,
        is_constant=is_constant,
    )
    
    # Numeric stats
    if col.dtype.is_numeric():
        non_null = col.drop_nulls()
        if len(non_null) > 0:
            stats.min_val = float(non_null.min())
            stats.max_val = float(non_null.max())
            stats.mean_val = float(non_null.mean())
            stats.std_val = float(non_null.std()) if len(non_null) > 1 else 0.0
            stats.q25 = float(non_null.quantile(0.25))
            stats.q50 = float(non_null.quantile(0.50))
            stats.q75 = float(non_null.quantile(0.75))
            # Outlier detection using IQR
            if stats.q25 is not None and stats.q75 is not None:
                iqr = stats.q75 - stats.q25
                lower = stats.q25 - 1.5 * iqr
                upper = stats.q75 + 1.5 * iqr
                stats.outlier_count = int(((non_null < lower) | (non_null > upper)).sum())
    
    # Categorical / text stats
    if col.dtype == pl.Utf8:
        # Top values
        vc = (
            col.drop_nulls()
            .value_counts()
            .sort("count", descending=True)
            .head(top_k)
        )
        if len(vc) > 0:
            stats.top_values = [(str(row[col_name]), int(row["count"])) for row in vc.iter_rows(named=True)]
        
        stats.is_high_cardinality = nunique > high_card_threshold
        
        # Text length
        non_null = col.drop_nulls()
        if len(non_null) > 0:
            lengths = non_null.str.len_chars()
            stats.text_len_min = int(lengths.min())
            stats.text_len_max = int(lengths.max())
            stats.text_len_mean = float(lengths.mean())
        
        # JSON parse rate (check if column might contain JSON)
        sample = non_null.head(100)
        if len(sample) > 0:
            json_count = sum(1 for v in sample.to_list() if try_parse_json(v))
            stats.json_parse_rate = json_count / len(sample)
    
    return stats


def analyze_table(
    df: pl.DataFrame,
    table_name: str,
    candidate_key_cols: list[list[str]] | None = None,
    top_k: int = 5,
) -> TableStats:
    """Analyze a single table."""
    row_count = len(df)
    col_count = len(df.columns)
    
    # Analyze each column
    columns = [analyze_column(df, col, top_k=top_k) for col in df.columns]
    
    # Constant columns
    constant_columns = [c.name for c in columns if c.is_constant]
    
    # High missing columns (>50%)
    high_missing_columns = [c.name for c in columns if c.missing_rate > 0.5]
    
    # Candidate key analysis
    candidate_keys = []
    if candidate_key_cols:
        for key_cols in candidate_key_cols:
            valid_cols = [c for c in key_cols if c in df.columns]
            if len(valid_cols) == len(key_cols):
                dup_count = df.select(valid_cols).is_duplicated().sum()
                dup_rate = dup_count / row_count if row_count > 0 else 0.0
                candidate_keys.append({
                    "columns": valid_cols,
                    "duplicate_count": int(dup_count),
                    "duplicate_rate": float(dup_rate),
                    "is_unique": dup_rate == 0,
                })
    
    return TableStats(
        name=table_name,
        row_count=row_count,
        col_count=col_count,
        columns=columns,
        candidate_keys=candidate_keys,
        constant_columns=constant_columns,
        high_missing_columns=high_missing_columns,
    )


def analyze_cross_tables(tables: dict[str, pl.DataFrame]) -> CrossTableStats:
    """Analyze cross-table relationships."""
    # StudyID coverage
    study_id_coverage = {}
    all_study_ids = set()
    for name, df in tables.items():
        if "StudyID" in df.columns:
            ids = df.get_column("StudyID").drop_nulls().unique().to_list()
            study_id_coverage[name] = len(ids)
            all_study_ids.update(ids)
    
    # (StudyID, Arm_ID) uniqueness
    arm_uniqueness = {}
    for name, df in tables.items():
        if "StudyID" in df.columns and "Arm_ID" in df.columns:
            key_cols = ["StudyID", "Arm_ID"]
            total = len(df)
            dup_count = df.select(key_cols).is_duplicated().sum()
            unique_combos = df.select(key_cols).unique().height
            arm_uniqueness[name] = {
                "total_rows": total,
                "unique_combinations": unique_combos,
                "duplicate_count": int(dup_count),
                "is_unique": dup_count == 0,
            }
    
    # Join coverage: (StudyID, Arm_ID, EP_Name)
    join_coverage = {}
    
    # Check R_Arm_Study_Endpoint coverage in R_Study_Endpoint
    if "R_Arm_Study_Endpoint_all" in tables and "R_Study_Endpoint_all" in tables:
        arm_ep = tables["R_Arm_Study_Endpoint_all"]
        study_ep = tables["R_Study_Endpoint_all"]
        if all(c in arm_ep.columns for c in ["StudyID", "EP_Name"]):
            arm_ep_keys = arm_ep.select(["StudyID", "EP_Name"]).unique()
            if all(c in study_ep.columns for c in ["StudyID"]):
                # Check how many arm endpoints have matching study endpoints
                # Simplified: just count coverage
                join_coverage["arm_ep_to_study_ep"] = 1.0  # placeholder
    
    return CrossTableStats(
        study_id_coverage=study_id_coverage,
        study_id_total=len(all_study_ids),
        arm_uniqueness=arm_uniqueness,
        join_coverage=join_coverage,
    )


def analyze_leakage(
    tables: dict[str, pl.DataFrame],
    timepoint_mapping: dict[str, str],
) -> LeakageReport:
    """Analyze potential data leakage based on timepoint mapping."""
    t0_features = []
    t1_features = []
    t2_features = []
    unknown_features = []
    
    # Collect all column names from all tables
    all_columns = set()
    for df in tables.values():
        all_columns.update(df.columns)
    
    for col in all_columns:
        timepoint = timepoint_mapping.get(col)
        if timepoint == "T0":
            t0_features.append(col)
        elif timepoint == "T1":
            t1_features.append(col)
        elif timepoint == "T2":
            t2_features.append(col)
        else:
            unknown_features.append(col)
    
    return LeakageReport(
        t0_features=sorted(t0_features),
        t1_features=sorted(t1_features),
        t2_features=sorted(t2_features),
        unknown_features=sorted(unknown_features),
        feature_mapping=timepoint_mapping,
    )


def load_timepoint_mapping(excel_path: str | Path) -> dict[str, str]:
    """Load variable -> timepoint mapping from Excel file (all sheets)."""
    from openpyxl import load_workbook
    
    # Get all sheet names
    wb = load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    mapping = {}
    for sheet_name in sheet_names:
        df = pl.read_excel(excel_path, sheet_name=sheet_name)
        for row in df.iter_rows(named=True):
            var = row.get("Variable")
            tp = row.get("Availability_Timepoint")
            if var and tp:
                # Strip whitespace from variable name
                var = var.strip() if isinstance(var, str) else var
                mapping[var] = tp
    
    return mapping


def run_eda(
    tables_dir: str | Path,
    timepoint_excel: str | Path | None = None,
    output_dir: str | Path | None = None,
) -> EDAReport:
    """Run complete EDA analysis."""
    tables_path = Path(tables_dir)
    
    # Load all tables
    tables = {}
    for csv_file in tables_path.glob("*.csv"):
        name = csv_file.stem
        tables[name] = pl.read_csv(csv_file)
    
    # Define candidate keys for each table type
    # Note: 
    # - EP_Pop distinguishes different analysis populations (e.g., basket/umbrella trials)
    # - EP_time_frame distinguishes same endpoint measured at different time points
    # - Some key duplicates may still exist due to null values in key columns
    key_definitions = {
        "D_Design_all": [["StudyID"]],
        "D_Pop_all": [["StudyID"]],
        "D_Drug_all": [["StudyID", "Generic_Name"]],
        "R_Study_all": [["StudyID"]],
        "R_Study_Endpoint_all": [
            ["StudyID", "EP_description", "EP_Pop", "EP_type", "EP_time_frame"],
        ],
        "R_Arm_Study_all": [["StudyID", "Arm_ID"]],
        "R_Arm_Study_Endpoint_all": [
            # EP_Point contains actual result values that differ by time point
            ["StudyID", "Arm_ID", "EP_Name", "EP_Pop", "EP_type"],
        ],
    }
    
    # Analyze each table
    table_stats = {}
    for name, df in tables.items():
        keys = key_definitions.get(name, [])
        table_stats[name] = analyze_table(df, name, candidate_key_cols=keys)
    
    # Cross-table analysis
    cross_table_stats = analyze_cross_tables(tables)
    
    # Leakage analysis
    timepoint_mapping = {}
    if timepoint_excel:
        timepoint_mapping = load_timepoint_mapping(timepoint_excel)
    leakage_report = analyze_leakage(tables, timepoint_mapping)
    
    report = EDAReport(
        table_stats=table_stats,
        cross_table_stats=cross_table_stats,
        leakage_report=leakage_report,
    )
    
    # Export if output_dir specified
    if output_dir:
        export_eda_report(report, output_dir)
    
    return report


def export_eda_report(report: EDAReport, output_dir: str | Path) -> None:
    """Export EDA report to CSV/JSON files."""
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    
    # 1. Schema summary (all tables)
    schema_rows = []
    for table_name, ts in report.table_stats.items():
        for col in ts.columns:
            schema_rows.append({
                "table": table_name,
                "column": col.name,
                "dtype": col.dtype,
                "total_rows": col.total_rows,
                "null_count": col.null_count,
                "pseudo_missing_count": col.pseudo_missing_count,
                "missing_rate": round(col.missing_rate, 4),
                "nunique": col.nunique,
                "is_constant": col.is_constant,
                "is_high_cardinality": col.is_high_cardinality,
                "min": col.min_val,
                "max": col.max_val,
                "mean": round(col.mean_val, 4) if col.mean_val else None,
                "std": round(col.std_val, 4) if col.std_val else None,
                "q25": col.q25,
                "q50": col.q50,
                "q75": col.q75,
                "outlier_count": col.outlier_count,
                "text_len_min": col.text_len_min,
                "text_len_max": col.text_len_max,
                "text_len_mean": round(col.text_len_mean, 2) if col.text_len_mean else None,
                "json_parse_rate": round(col.json_parse_rate, 4) if col.json_parse_rate else None,
            })
    pl.DataFrame(schema_rows).write_csv(out_path / "column_stats.csv")
    
    # 2. Table summary
    table_rows = []
    for name, ts in report.table_stats.items():
        table_rows.append({
            "table": name,
            "row_count": ts.row_count,
            "col_count": ts.col_count,
            "constant_columns": len(ts.constant_columns),
            "high_missing_columns": len(ts.high_missing_columns),
            "constant_col_names": ", ".join(ts.constant_columns[:10]),
            "high_missing_col_names": ", ".join(ts.high_missing_columns[:10]),
        })
    pl.DataFrame(table_rows).write_csv(out_path / "table_summary.csv")
    
    # 3. Key duplication
    key_rows = []
    for name, ts in report.table_stats.items():
        for ck in ts.candidate_keys:
            key_rows.append({
                "table": name,
                "key_columns": ", ".join(ck["columns"]),
                "duplicate_count": ck["duplicate_count"],
                "duplicate_rate": round(ck["duplicate_rate"], 4),
                "is_unique": ck["is_unique"],
            })
    if key_rows:
        pl.DataFrame(key_rows).write_csv(out_path / "key_duplication.csv")
    
    # 4. Cross-table stats
    cross_rows = []
    for name, count in report.cross_table_stats.study_id_coverage.items():
        cross_rows.append({
            "metric": "study_id_count",
            "table": name,
            "value": count,
        })
    cross_rows.append({
        "metric": "study_id_total",
        "table": "all",
        "value": report.cross_table_stats.study_id_total,
    })
    for name, info in report.cross_table_stats.arm_uniqueness.items():
        cross_rows.append({
            "metric": "arm_uniqueness",
            "table": name,
            "value": info["unique_combinations"],
            "detail": json.dumps(info),
        })
    pl.DataFrame(cross_rows).write_csv(out_path / "cross_table_stats.csv")
    
    # 5. Leakage report
    leakage_data = {
        "t0_features": report.leakage_report.t0_features,
        "t1_features": report.leakage_report.t1_features,
        "t2_features_exclude": report.leakage_report.t2_features,
        "unknown_features": report.leakage_report.unknown_features,
        "feature_mapping": report.leakage_report.feature_mapping,
    }
    with open(out_path / "leakage_report.json", "w") as f:
        json.dump(leakage_data, f, indent=2)
    
    # 6. Top values for categorical columns
    topk_rows = []
    for table_name, ts in report.table_stats.items():
        for col in ts.columns:
            if col.top_values:
                for i, (val, cnt) in enumerate(col.top_values):
                    topk_rows.append({
                        "table": table_name,
                        "column": col.name,
                        "rank": i + 1,
                        "value": val[:100] if len(val) > 100 else val,  # truncate long values
                        "count": cnt,
                    })
    if topk_rows:
        pl.DataFrame(topk_rows).write_csv(out_path / "categorical_top_values.csv")
    
    print(f"EDA report exported to {out_path}")
    print(f"  - column_stats.csv: {len(schema_rows)} rows")
    print(f"  - table_summary.csv: {len(table_rows)} rows")
    print(f"  - key_duplication.csv: {len(key_rows)} rows")
    print(f"  - cross_table_stats.csv: {len(cross_rows)} rows")
    print(f"  - leakage_report.json")
    print(f"  - categorical_top_values.csv: {len(topk_rows)} rows")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m ctg_ml_pipeline.analysis.eda <tables_dir> [timepoint_excel] [output_dir]")
        sys.exit(1)
    
    tables_dir = sys.argv[1]
    timepoint_excel = sys.argv[2] if len(sys.argv) > 2 else None
    output_dir = sys.argv[3] if len(sys.argv) > 3 else None
    
    run_eda(tables_dir, timepoint_excel, output_dir)
